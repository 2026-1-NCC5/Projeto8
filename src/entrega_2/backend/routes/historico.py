"""
Rotas do histórico de alimentos: CRUD + paginação + exportação Excel.
"""

import io
import math
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from database import get_db
from models.usuario import Usuario
from models.equipe import Equipe, EquipeMembro, EquipeMentor
from models.semestre import Semestre
from models.historico_alimento import HistoricoAlimento
from schemas.historico_alimento import (
    HistoricoAlimentoCreate, HistoricoAlimentoUpdate,
    HistoricoAlimentoResponse, HistoricoPaginado,
)
from auth.dependencies import get_current_user, require_admin

router = APIRouter(tags=["Histórico de Alimentos"])


# ── GET /equipes/:id/historico ──────────────────
@router.get("/api/equipes/{equipe_id}/historico", response_model=HistoricoPaginado)
def listar_historico(
    equipe_id: int,
    semestre_id: int = Query(None, description="Filtrar por semestre"),
    pagina: int = Query(1, ge=1),
    por_pagina: int = Query(5, ge=1, le=50),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """Lista o histórico de alimentos de uma equipe (paginado)."""
    equipe = db.query(Equipe).filter(Equipe.id == equipe_id).first()
    if not equipe:
        raise HTTPException(status_code=404, detail="Equipe não encontrada.")

    query = db.query(HistoricoAlimento).filter(HistoricoAlimento.equipe_id == equipe_id)
    total = query.count()
    items = (
        query.order_by(HistoricoAlimento.data.desc())
        .offset((pagina - 1) * por_pagina)
        .limit(por_pagina)
        .all()
    )

    return HistoricoPaginado(
        items=items,
        total=total,
        pagina=pagina,
        por_pagina=por_pagina,
        total_paginas=math.ceil(total / por_pagina) if total > 0 else 1,
    )


# ── POST /equipes/:id/historico ─────────────────
@router.post(
    "/api/equipes/{equipe_id}/historico",
    response_model=HistoricoAlimentoResponse,
    status_code=201,
)
def criar_registro(
    equipe_id: int,
    dados: HistoricoAlimentoCreate,
    db: Session = Depends(get_db),
    admin: Usuario = Depends(require_admin),
):
    """Adiciona um registro de alimento (somente admin)."""
    if not db.query(Equipe).filter(Equipe.id == equipe_id).first():
        raise HTTPException(status_code=404, detail="Equipe não encontrada.")

    registro = HistoricoAlimento(
        equipe_id=equipe_id,
        data=dados.data,
        item=dados.item,
        quantidade=dados.quantidade,
        unidade=dados.unidade,
        status=dados.status,
    )
    db.add(registro)
    db.commit()
    db.refresh(registro)
    return registro


# ── PUT /historico/:id ──────────────────────────
@router.put("/api/historico/{registro_id}", response_model=HistoricoAlimentoResponse)
def editar_registro(
    registro_id: int,
    dados: HistoricoAlimentoUpdate,
    db: Session = Depends(get_db),
    admin: Usuario = Depends(require_admin),
):
    """Edita um registro de alimento (somente admin)."""
    reg = db.query(HistoricoAlimento).filter(HistoricoAlimento.id == registro_id).first()
    if not reg:
        raise HTTPException(status_code=404, detail="Registro não encontrado.")

    for field in ["data", "item", "quantidade", "unidade", "status"]:
        val = getattr(dados, field, None)
        if val is not None:
            setattr(reg, field, val)

    db.commit()
    db.refresh(reg)
    return reg


# ── DELETE /historico/:id ───────────────────────
@router.delete("/api/historico/{registro_id}", status_code=204)
def excluir_registro(
    registro_id: int,
    db: Session = Depends(get_db),
    admin: Usuario = Depends(require_admin),
):
    """Exclui um registro de alimento (somente admin)."""
    reg = db.query(HistoricoAlimento).filter(HistoricoAlimento.id == registro_id).first()
    if not reg:
        raise HTTPException(status_code=404, detail="Registro não encontrado.")
    db.delete(reg)
    db.commit()


# ── GET /equipes/:id/exportar ───────────────────
@router.get("/api/equipes/{equipe_id}/exportar")
def exportar_relatorio(
    equipe_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """Exporta relatório da equipe em Excel (.xlsx)."""
    equipe = db.query(Equipe).filter(Equipe.id == equipe_id).first()
    if not equipe:
        raise HTTPException(status_code=404, detail="Equipe não encontrada.")

    semestre = db.query(Semestre).filter(Semestre.id == equipe.semestre_id).first()

    # Dados
    membros = (
        db.query(Usuario)
        .join(EquipeMembro, EquipeMembro.usuario_id == Usuario.id)
        .filter(EquipeMembro.equipe_id == equipe_id)
        .all()
    )
    mentores = (
        db.query(Usuario)
        .join(EquipeMentor, EquipeMentor.usuario_id == Usuario.id)
        .filter(EquipeMentor.equipe_id == equipe_id)
        .all()
    )
    historico = (
        db.query(HistoricoAlimento)
        .filter(HistoricoAlimento.equipe_id == equipe_id)
        .order_by(HistoricoAlimento.data.desc())
        .all()
    )

    total_itens = sum(h.quantidade for h in historico)
    total_registros = len(historico)

    # Criar Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório da Equipe"

    green_fill = PatternFill(start_color="006E2F", end_color="006E2F", fill_type="solid")
    light_fill = PatternFill(start_color="A4F1B2", end_color="A4F1B2", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    title_font = Font(name="Arial", bold=True, size=14)
    bold = Font(name="Arial", bold=True, size=11)
    normal = Font(name="Arial", size=11)
    thin = Side(style="thin", color="BCCBB9")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    row = 1
    # Título
    ws.merge_cells("A1:E1")
    ws["A1"] = f"Relatório - Equipe {equipe.nome}"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    row = 3
    # Info Semestre
    ws[f"A{row}"] = "Semestre:"
    ws[f"A{row}"].font = bold
    ws[f"B{row}"] = semestre.nome if semestre else "N/A"
    row += 1
    ws[f"A{row}"] = "Status:"
    ws[f"A{row}"].font = bold
    ws[f"B{row}"] = semestre.status.upper() if semestre else "N/A"
    row += 1
    ws[f"A{row}"] = "Gerado em:"
    ws[f"A{row}"].font = bold
    ws[f"B{row}"] = datetime.now().strftime("%d/%m/%Y %H:%M")

    # Big Numbers
    row += 2
    ws.merge_cells(f"A{row}:B{row}")
    ws[f"A{row}"] = "RESUMO"
    ws[f"A{row}"].font = Font(name="Arial", bold=True, size=13, color="006E2F")
    row += 1
    ws[f"A{row}"] = "Total Arrecadado:"
    ws[f"A{row}"].font = bold
    ws[f"B{row}"] = f"{total_itens:,} itens".replace(",", ".")
    ws[f"B{row}"].font = Font(name="Arial", bold=True, size=14, color="006E2F")
    row += 1
    ws[f"A{row}"] = "Total Registros:"
    ws[f"A{row}"].font = bold
    ws[f"B{row}"] = str(total_registros)

    # Mentores
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "MENTORES"
    ws[f"A{row}"].font = header_font
    ws[f"A{row}"].fill = green_fill
    for c in ["B", "C", "D", "E"]:
        ws[f"{c}{row}"].fill = green_fill
    row += 1
    ws[f"A{row}"] = "Nome"
    ws[f"A{row}"].font = bold
    ws[f"A{row}"].fill = light_fill
    ws[f"B{row}"] = "E-mail"
    ws[f"B{row}"].font = bold
    ws[f"B{row}"].fill = light_fill
    row += 1
    for mentor in mentores:
        ws[f"A{row}"] = mentor.nome
        ws[f"B{row}"] = mentor.email
        row += 1

    # Membros
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "MEMBROS (ALUNOS)"
    ws[f"A{row}"].font = header_font
    ws[f"A{row}"].fill = green_fill
    for c in ["B", "C", "D", "E"]:
        ws[f"{c}{row}"].fill = green_fill
    row += 1
    for col, label in [("A", "Nome"), ("B", "Curso"), ("C", "RA"), ("D", "E-mail")]:
        ws[f"{col}{row}"] = label
        ws[f"{col}{row}"].font = bold
        ws[f"{col}{row}"].fill = light_fill
    row += 1
    for membro in membros:
        ws[f"A{row}"] = membro.nome
        ws[f"B{row}"] = membro.curso or "—"
        ws[f"C{row}"] = membro.ra or "—"
        ws[f"D{row}"] = membro.email
        row += 1

    # Histórico
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "HISTÓRICO DE ALIMENTOS"
    ws[f"A{row}"].font = header_font
    ws[f"A{row}"].fill = green_fill
    for c in ["B", "C", "D", "E"]:
        ws[f"{c}{row}"].fill = green_fill
    row += 1
    for col, label in [("A", "Data"), ("B", "Item"), ("C", "Quantidade"), ("D", "Unidade"), ("E", "Status")]:
        ws[f"{col}{row}"] = label
        ws[f"{col}{row}"].font = bold
        ws[f"{col}{row}"].fill = light_fill
    row += 1
    for h in historico:
        ws[f"A{row}"] = h.data.strftime("%d/%m/%Y")
        ws[f"B{row}"] = h.item
        ws[f"C{row}"] = h.quantidade
        ws[f"D{row}"] = h.unidade
        ws[f"E{row}"] = h.status.capitalize()
        row += 1

    # Ajustar larguras
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 15

    # Salvar em memória
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"relatorio_equipe_{equipe.nome}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
